from tkinter import *
from tkinter import filedialog
from pathlib import PureWindowsPath
import pandas as pd
import numpy as np
np.seterr(divide='ignore', invalid='ignore')
import matplotlib.pyplot as plt
#plt.style.use('fivethirtyeight')
from matplotlib.figure import Figure 
from matplotlib.pyplot import imshow
import os
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import tkinter.ttk as ttk
import seaborn as sns
sns.set_context("paper")
from tkinter.messagebox import *
import statistics as stq
from sklearn.model_selection import learning_curve
from collections import OrderedDict
#from time import time
import time as tm
from random import random
import datetime
import locale
locale.setlocale(locale.LC_TIME,'')


fenetre = Tk()

w=fenetre.winfo_screenwidth()
h=fenetre.winfo_screenheight()
fenetre.title("PREVISION DE VENTE (PDV)")
fenetre.minsize(w,h)
fenetre.maxsize(w,h)
fenetre.iconbitmap('icon/p.ico')

###############################################################################
#################### CONNEXION AU BASE DE DONNEE ##############################
import sqlite3
#creation de base de donnee
conn = sqlite3.connect('prevision.db')
c = conn.cursor()
#Creation des table 

c.execute('''CREATE TABLE IF NOT EXISTS login (username text, password text)''')

c.execute('''CREATE TABLE IF NOT EXISTS historique
			 (Nom text, Date_Creation date, Type_data text, Type_traitement text, Score Real, Modele INTEGER, Target text,Prediction real)''')

#Insertion login

#c.execute("INSERT INTO login VALUES ('ADMIN','admin')")

# Save (commit) the changes
conn.commit()

########################## FIN #################################################
################################################################################


################################################################################
##################### DEBUT INTERFACE GENERALE DE PDV ##########################
################################################################################

def regression():

	from sklearn.pipeline import make_pipeline
	from sklearn.feature_selection import SelectKBest, f_classif
	from sklearn.preprocessing import PolynomialFeatures, StandardScaler, OneHotEncoder
	from sklearn.metrics import r2_score,mean_absolute_error,mean_squared_error,median_absolute_error,confusion_matrix
	from sklearn.gaussian_process.kernels import RBF, ConstantKernel,WhiteKernel,DotProduct
	from sklearn.model_selection import learning_curve,cross_val_score

	from sklearn.linear_model import SGDRegressor
	from sklearn.linear_model import Ridge
	from sklearn.svm import SVR
	from sklearn.neighbors import KNeighborsRegressor
	from sklearn.gaussian_process import  GaussianProcessRegressor
	from sklearn.tree import  DecisionTreeRegressor
	from sklearn.neural_network import MLPRegressor
	from sklearn.ensemble import RandomForestRegressor
	from sklearn.ensemble import ExtraTreesRegressor
	from sklearn.ensemble import GradientBoostingRegressor
	from sklearn.ensemble import AdaBoostRegressor

	global Ridge,SGD,SVM,KNN,Tree,Extratree,RNA,RForest,GBoosting,AdaBoost,Gaussian

	import matplotlib.pyplot as plt

	kernel_gp = DotProduct() + WhiteKernel()

	preprocessor=make_pipeline(PolynomialFeatures(2, include_bias=False),StandardScaler(), SelectKBest(f_classif,k='all'))

	Ridge=make_pipeline(preprocessor, Ridge(random_state=0))
	SGD=make_pipeline(preprocessor,SGDRegressor(loss='squared_loss',random_state=0,max_iter=1000, tol=1e-3))
	SVM=make_pipeline(preprocessor,SVR(kernel='linear'))

	KNN=make_pipeline(preprocessor, KNeighborsRegressor(n_neighbors=2,weights='uniform',leaf_size=30,algorithm='auto'))
	Tree=make_pipeline(preprocessor,DecisionTreeRegressor(max_features=5,max_depth=4, random_state=0,min_samples_split=5))#,min_samples_split=5
	Extratree=make_pipeline(preprocessor,ExtraTreesRegressor(max_features=5,random_state=0,max_depth=4))

	RNA=make_pipeline(preprocessor, MLPRegressor(solver='lbfgs',hidden_layer_sizes=(8,),random_state=0,max_iter=1000))
	RForest=make_pipeline(preprocessor,RandomForestRegressor(random_state=0))
	GBoosting=make_pipeline(preprocessor,GradientBoostingRegressor(random_state=0))
	AdaBoost=make_pipeline(preprocessor, AdaBoostRegressor(random_state=0,n_estimators=40, loss='linear',learning_rate=1.0))
	Gaussian=make_pipeline(preprocessor,GaussianProcessRegressor(kernel=kernel_gp,n_restarts_optimizer=0, normalize_y=True, alpha=.5))#1e-20

	global liste_pred, nom_model, list_score_train, list_score_val
	global score_ridge,score_sgd,score_svm,score_knn,score_tree,score_extratree,score_rna,score_rforest,score_gboosting,score_adaboost,score_gaussian
	global t_ridge,t_sgd,t_svm,t_knn,t_tree,t_extratree,t_rna,t_rforest,t_gboosting,t_adaboost,t_gaussian
	global pred_ridge,pred_sgd,pred_svm,pred_knn,pred_tree,pred_extratree,pred_rna,pred_rforest,pred_gboosting,pred_adaboost,pred_gaussian


	if type_dataset==liste_dataset[0]:

		def evaluation(model):
			t0=tm.time()
			model.fit(X_train,y_train)
			prediction=model.predict(X_test)
			#print(prediction)
			y_pred=prediction.mean().round(4)
			train_score=model.score(X_train,y_train).round(4)
			#train_score=train_scr.round(4)

			test_score=model.score(X_test,y_test).round(4)

			t1=tm.time()
			duree=t1-t0

			global dict_metric,erreur_hist

			MSE=mean_absolute_error(y_test,prediction)
			MAE=mean_squared_error(y_test,prediction)
			RMSE=np.sqrt(mean_absolute_error(y_test,prediction))
			r2=r2_score(y_test,prediction)

			dict_metric={'MSE':MSE,'MAE':MAE,'RMSE':RMSE,'R**2':r2}

			erreur_hist=np.abs(y_test-prediction)

			return y_pred, train_score, duree, test_score

		pred_ridge, score_ridge,t_ridge,score_ridge_val=evaluation(Ridge)

		pred_sgd, score_sgd,t_sgd, score_sgd_val=evaluation(SGD)

		pred_svm, score_svm,t_svm,score_svm_val=evaluation(SVM)

		pred_knn, score_knn,t_knn, score_knn_val=evaluation(KNN)

		pred_tree, score_tree,t_tree, score_tree_val=evaluation(Tree)
	
		pred_extratree, score_extratree,t_extratree, score_extratree_val=evaluation(Extratree)
	
		pred_rna, score_rna,t_rna, score_rna_val=evaluation(RNA)
	
		pred_rforest, score_rforest,t_rforest, score_rforest_val=evaluation(RForest)
	
		pred_gboosting, score_gboosting,t_gboosting, score_gboosting_val=evaluation(GBoosting)
	
		pred_adaboost, score_adaboost,t_adaboost, score_adaboost_val=evaluation(AdaBoost)
	
		pred_gaussian, score_gaussian,t_gaussian, score_gaussian_val=evaluation(Gaussian)


		nom_model=('Ridge','SGD','SVM','KNN','Tree','Extratree','RNA','RForest', 'GBoosting','AdaBoost','Gaussian')
		liste_pred=[pred_ridge,pred_sgd,pred_svm,pred_knn,pred_tree,pred_extratree,pred_rna,pred_rforest,pred_gboosting,pred_adaboost,pred_gaussian]
		
		list_score_train=[score_ridge,score_sgd,score_svm,score_knn,score_tree,score_extratree,score_rna,score_rforest,score_gboosting,score_adaboost,score_gaussian]
		list_score_val=[score_ridge_val,score_sgd_val,score_svm_val,score_knn_val,score_tree_val,score_extratree_val,score_rna_val,score_rforest_val,score_gboosting_val,score_adaboost_val,score_gaussian_val]
		

	if type_dataset==liste_dataset[1]: #serie temporelle
		def evaluation(model):
			t0=tm.time()
			model.fit(X_train,y_train)
			prediction=model.predict(X_test)
			y_pred=prediction.mean().round(4)+derniere_val
			train_score=model.score(X_train,y_train)
			#train_score=train_scr.round(4)

			test_score=model.score(X_test,y_test)


			t1=tm.time()
			duree=t1-t0

			global dict_metric, erreur_hist

			MSE=mean_absolute_error(y_test,prediction)
			MAE=mean_squared_error(y_test,prediction)
			RMSE=np.sqrt(mean_absolute_error(y_test,prediction))
			r2=r2_score(y_test,prediction)

			dict_metric={'MSE':MSE,'MAE':MAE,'RMSE':RMSE,'R2':r2}

			erreur_hist=np.abs(y_test-prediction)

			return y_pred, train_score, duree, test_score


		pred_ridge, score_ridge,t_ridge,score_ridge_val=evaluation(Ridge)

		pred_sgd, score_sgd,t_sgd, score_sgd_val=evaluation(SGD)

		pred_svm, score_svm,t_svm,score_svm_val=evaluation(SVM)

		pred_knn, score_knn,t_knn, score_knn_val=evaluation(KNN)

		pred_tree, score_tree,t_tree, score_tree_val=evaluation(Tree)

		pred_extratree, score_extratree,t_extratree, score_extratree_val=evaluation(Extratree)

		pred_rna, score_rna,t_rna, score_rna_val=evaluation(RNA)

		pred_rforest, score_rforest,t_rforest, score_rforest_val=evaluation(RForest)

		pred_gboosting, score_gboosting,t_gboosting, score_gboosting_val=evaluation(GBoosting)

		pred_adaboost, score_adaboost,t_adaboost, score_adaboost_val=evaluation(AdaBoost)

		pred_gaussian, score_gaussian,t_gaussian, score_gaussian_val=evaluation(Gaussian)


		nom_model=('Ridge','SGD','SVM','KNN','Tree','Extratree','RNA','RForest', 'GBoosting','AdaBoost','Gaussian')
		liste_pred=[pred_ridge[0],pred_sgd[0],pred_svm[0],pred_knn[0],pred_tree[0],pred_extratree[0],pred_rna[0],pred_rforest[0],pred_gboosting[0],pred_adaboost[0],pred_gaussian[0]]
		list_score_train=[score_ridge,score_sgd,score_svm,score_knn,score_tree,score_extratree,score_rna,score_rforest,score_gboosting,score_adaboost,score_gaussian]
		list_score_val=[score_ridge_val,score_sgd_val,score_svm_val,score_knn_val,score_tree_val,score_extratree_val,score_rna_val,score_rforest_val,score_gboosting_val,score_adaboost_val,score_gaussian_val]


def classification():


	global Ridge,SGD,SVM,KNN,Tree,Extratree,RNA,RForest,GBoosting,AdaBoost,Gaussian, Bayes

	import matplotlib.pyplot as plt

	from sklearn.pipeline import make_pipeline
	from sklearn.feature_selection import SelectKBest, f_classif
	from sklearn.preprocessing import PolynomialFeatures, StandardScaler, OneHotEncoder
	from sklearn.metrics import r2_score,confusion_matrix,classification_report
	from sklearn.gaussian_process.kernels import RBF, ConstantKernel,WhiteKernel,DotProduct

	from sklearn.linear_model import SGDClassifier
	from sklearn.linear_model import RidgeClassifier
	from sklearn.svm import SVC
	from sklearn.neighbors import KNeighborsClassifier
	from sklearn.gaussian_process import  GaussianProcessClassifier
	from sklearn.tree import  DecisionTreeClassifier
	from sklearn.neural_network import MLPClassifier
	from sklearn.ensemble import RandomForestClassifier
	from sklearn.ensemble import ExtraTreesClassifier
	from sklearn.ensemble import GradientBoostingClassifier
	from sklearn.ensemble import AdaBoostClassifier
	from sklearn.model_selection import learning_curve,cross_val_score
	from sklearn.naive_bayes import GaussianNB
	from sklearn.metrics import accuracy_score, f1_score, precision_score, recall_score, classification_report, confusion_matrix

	from time import time

	import matplotlib.pyplot as plt

	kernel_gp = DotProduct() + WhiteKernel()

	preprocessor=make_pipeline(PolynomialFeatures(2, include_bias=False),StandardScaler(), SelectKBest(f_classif,k=6))

	Ridge=make_pipeline(preprocessor, RidgeClassifier(random_state=0))
	SGD=make_pipeline(preprocessor,SGDClassifier(loss='squared_loss',random_state=0,max_iter=1000, tol=1e-3))
	SVM=make_pipeline(preprocessor,SVC(kernel='linear'))

	KNN=make_pipeline(preprocessor, KNeighborsClassifier(n_neighbors=2,weights='uniform',leaf_size=30,algorithm='auto'))
	Tree=make_pipeline(preprocessor,DecisionTreeClassifier(max_features=5,max_depth=4, random_state=0,min_samples_split=5))#,min_samples_split=5
	Extratree=make_pipeline(preprocessor,ExtraTreesClassifier(max_features=5,random_state=0,max_depth=4))

	RNA=make_pipeline(preprocessor, MLPClassifier(solver='lbfgs',hidden_layer_sizes=(8,),random_state=0,max_iter=1000))
	RForest=make_pipeline(preprocessor,RandomForestClassifier(random_state=0))
	GBoosting=make_pipeline(preprocessor,GradientBoostingClassifier(random_state=0))
	AdaBoost=make_pipeline(preprocessor, AdaBoostClassifier(random_state=0,n_estimators=40,learning_rate=1.0))
	Gaussian=make_pipeline(preprocessor,GaussianProcessClassifier(kernel=kernel_gp,n_restarts_optimizer=0))#1e-20
	Bayes=make_pipeline(preprocessor, GaussianNB())

	def evaluation(model):
		global prediction
		t0=tm.time()
		model.fit(X_train,y_train)

		prediction=model.predict(X_test)
		#prediction=model.predict_proba(X_test)
		y_pred=prediction.mean().round(4)
		train_score=model.score(X_train,y_train).round(4)
		#train_score=train_scr.round(4)

		test_score=model.score(X_test,y_test).round(4)

		t1=tm.time()
		duree=t1-t0
		
		Acc=accuracy_score(y_test, prediction)
		F1=f1_score(y_test, prediction,average='weighted')
		Rec=recall_score(y_test, prediction,average='weighted')
		prec=precision_score(y_test, prediction,average='weighted')
		
		global mat_conf
		mat_conf=confusion_matrix(y_test, prediction)

		global dict_metric
		dict_metric={'Acc': Acc,'F1':F1,'Rec':Rec,'Prec':prec}



		return y_pred, train_score, duree, test_score

	global score_ridge,score_sgd,score_svm,score_knn,score_tree,score_extratree,score_rna,score_rforest,score_gboosting,score_adaboost,score_gaussian, score_bayes
	global t_ridge,t_sgd,t_svm,t_knn,t_tree,t_extratree,t_rna,t_rforest,t_gboosting,t_adaboost,t_gaussian, t_bayes
	global pred_ridge,pred_sgd,pred_svm,pred_knn,pred_tree,pred_extratree,pred_rna,pred_rforest,pred_gboosting,pred_adaboost,pred_gaussian, pred_bayes

	pred_ridge, score_ridge,t_ridge,score_ridge_val=evaluation(Ridge)

	pred_sgd, score_sgd,t_sgd, score_sgd_val=evaluation(SGD)

	pred_svm, score_svm,t_svm,score_svm_val=evaluation(SVM)

	pred_knn, score_knn,t_knn, score_knn_val=evaluation(KNN)

	pred_tree, score_tree,t_tree, score_tree_val=evaluation(Tree)

	pred_extratree, score_extratree,t_extratree, score_extratree_val=evaluation(Extratree)

	pred_rna, score_rna,t_rna, score_rna_val=evaluation(RNA)

	pred_rforest, score_rforest,t_rforest, score_rforest_val=evaluation(RForest)

	pred_gboosting, score_gboosting,t_gboosting, score_gboosting_val=evaluation(GBoosting)

	pred_adaboost, score_adaboost,t_adaboost, score_adaboost_val=evaluation(AdaBoost)

	pred_gaussian, score_gaussian,t_gaussian, score_gaussian_val=evaluation(Gaussian)

	pred_bayes, score_bayes,t_bayes, score_bayes_val=evaluation(Bayes)

	global liste_pred, nom_model, list_score_train, list_score_val

	nom_model=('Ridge','SGD','SVM','KNN','Tree','Extratree','RNA','RForest', 'GBoosting','AdaBoost','Gaussian', 'NBayes')
	liste_pred=[pred_ridge,pred_sgd,pred_svm,pred_knn,pred_tree,pred_extratree,pred_rna,pred_rforest,pred_gboosting,pred_adaboost,pred_gaussian, pred_bayes]
	list_score_train=[score_ridge,score_sgd,score_svm,score_knn,score_tree,score_extratree,score_rna,score_rforest,score_gboosting,score_adaboost,score_gaussian,score_bayes]
	list_score_val=[score_ridge_val,score_sgd_val,score_svm_val,score_knn_val,score_tree_val,score_extratree_val,score_rna_val,score_rforest_val,score_gboosting_val,score_adaboost_val,score_gaussian_val,score_bayes_val]


def login():

	def connecter():
		
		#Login
		global user_name

		sql="Select * from login where username=? and password=?"
		c.execute(sql, (nom_societe.get(), pwd_societe.get()))

		res=c.fetchall()

		if res:
			importer()
			societe.set(nom_societe.get())
			user_name=nom_societe.get()
			login.destroy()
		else:
			showwarning('Erreur login', 'Mot de passe incorrect!')


	#interface login
			
	login=Toplevel(fenetre, width=200, height=200)
	login.title("Login")
	
	nom_societe=StringVar()
	pwd_societe=StringVar()
	
	lab_user=Label(login, text="Utilisateur :")
	lab_user.grid(row=0, column=0)
	user=Entry(login, width=20, textvariable=nom_societe, font='times 14')
	user.grid(row=0, column=1)
	lab_mdp=Label(login, text="Mot de passe :")
	lab_mdp.grid(row=1, column=0)
	pwd=Entry(login, width=20, textvariable=pwd_societe , show='*', font='times 14')
	pwd.grid(row=1, column=1)
	btn_connect=Button(login, width=25,text="Connecter",bg="blue",fg="white", command=connecter)
	btn_connect.grid(row=4, column=1)

	#fin de l'interface login


def importer():
	global data,liste_colonne, name_fic,col_lig
	
	type_file={("Fichier CSV","*.csv"),
			   ("Fichier EXCEL",("*.xlsx","*.xls"))
		}
	
	fichier =  filedialog.askopenfilename(initialdir =r"C:\Desktop",title = "Selectioner un fichier",filetypes = type_file)
	name_fic = PureWindowsPath(fichier).name

	#tester l'extension de fichier
	
	global extension
	extension=fichier.split('.')[-1]
	
	if extension=='csv':
		data=pd.read_csv(fichier)
	else:
		data=pd.read_excel(fichier)
	#encodage de donnee
	def encodage(df):
		for col in df.select_dtypes('object').columns:
			df.loc[:,col]=df[col].astype('category').cat.codes
		return df

	if fichier:
		col_lig=data.shape
		liste_colonne=list(data)
		dataset()
		data=data.dropna(axis=0)
		data=encodage(data)
	else:
		showwarning('Erreur d importation', 'Votre fichier est corrompu!')

###execution du modele
def executer():

	def progress_bar_func():
		global num

		num = 1
		f_progress.after(100, update_progress_bar)

	def update_progress_bar():
		global num

		if num <= brt:
			percentage = round(num/brt*100)  # Calculate percentage.
			progress['value'] = num
			num += 10
			if num > brt:
				showinfo(title="Information", message='Prédiction terminée!\n\n Vous pouvez accéder au menu "Générale"')
			else:
				f_progress.after(100, update_progress_bar)

	global mtrq,type_score, X_train,X_test,y_test,y_train

	try:
		if type_dataset==liste_dataset[0]:
			if type_entrainement==liste_entrainement[0]:
				regression()
				mtrq='MSE'
				type_score='neg_mean_squared_error'
			if type_entrainement==liste_entrainement[1]:
				classification()
				mtrq='Accuracy'  
				type_score='accuracy' 
		if type_dataset==liste_dataset[1]:
			regression()
			mtrq='MSE'
			type_score='neg_mean_squared_error'
	except:
		showwarning('Erreur', "Votre données n'est pas définie!")

	progress_bar_func()

def dataset():


	def serieTemporelle(data):
		from statsmodels.tsa.tsatools import lagmat
		from statsmodels.tsa.arima_model import ARIMA

		global df
		df=data.copy()
		
		X=df[type_index]
		y=df[type_target]

		# faire ne differenciation
		df["diff"] = np.nan
		df.loc[1:, "diff"] = (df.iloc[1:, 1].values - df.iloc[:len(df)-1, 1].values)

		global derniere_val,res_arima, forecasting

		derniere_val=df.iloc[len(df)-1:len(df), 1].values
		#model ARIMA

		forecasting=df[type_target].iloc[1:]
		try:
			arima_mod = ARIMA(forecasting, order=(1, 1, 1))
			res_arima = arima_mod.fit()
			#prevision_ts=res_arima.forecast(steps=0)[0]

			#print("Prevision forecasting : ",prevision_ts)

		except:
			showwarning('Erreur', 'Erreur de forecasting!')
		
		#On créé la matrice avec les séries décalées.
		
		lag = 8
		X = lagmat(df["diff"], lag)
		lagged = df.copy()
		for c in range(1,lag+1):
			lagged["lag%d" % c] = X[:, c-1]

		#decoupe non aleatoire(serie temporelle) train/test
		xc = ["lag%d" % i for i in range(1,lag+1)]
		split = 0.66
		isplit = int(len(lagged) * split)
		xt = lagged[10:][xc]
		yt = lagged[10:]["diff"]

		X_train, y_train, X_test, y_test = xt[:isplit], yt[:isplit], xt[isplit:], yt[isplit:]


		return X_train,y_train,X_test,y_test

	#selection de type de donne
	
	def select_dataset(event=None):
		global type_dataset
		if event:
			type_dataset=event.widget.get()
			if type_dataset==liste_dataset[0]:
				cb_index.set('Par défaut')
				cb_index.configure(state='disable')
				cb_entrainement.configure(state='normal')
				cb_entrainement.set('Sélectionner')

			if type_dataset==liste_dataset[1]:
				cb_index.configure(state='normal')
				cb_index.set('Sélectionner')
				cb_entrainement.set('Régression')
				cb_entrainement.configure(state='disable')

			return type_dataset, data

	#selection de type d'entrainement
		
	def select_entrainement(event=None):
		global type_entrainement
		if event:
			type_entrainement=event.widget.get()
			
			return type_entrainement
		
	#selection de index pour le type de donnee serie temporelle

	def select_index(event=None):
		global type_index
		if event:
			type_index=event.widget.get()

	#selection de variable de target
		
	def select_target(event=None):
		
		from sklearn.model_selection import train_test_split

		global type_target, testset
		global X_train
		global y_train
		global X_test
		global y_test
		if event:
			type_target=event.widget.get()

			if type_dataset==liste_dataset[0]:
				#creation de trainset et testset
				
				trainset, testset=train_test_split(data, test_size=0.2, random_state=2)
				
				X_train=trainset.drop(type_target, axis=1)
				y_train=trainset[type_target]

				X_test=testset.drop(type_target, axis=1)
				y_test=testset[type_target]

			if type_dataset==liste_dataset[1]:

				X_train,y_train,X_test,y_test=serieTemporelle(data)

			showinfo(title="Indication", message="Cliquez sur le bouton </Prédire en-tete")

	def myfunction(event):
		canvas.configure(scrollregion=canvas.bbox("all"),width=w-150,height=h-100)

	
	frame_data.pack(side=TOP, padx=0)
	mb.configure(state='normal')
	frame_accueil.pack_forget()
	frame_general.pack_forget()
	frame_model.pack_forget()
	frame_prevision.pack_forget()
	frame_features.pack_forget()
	frame_score.pack_forget()
	frame_historique.pack_forget()  
	btn_predire.configure(state='normal')
	
	# create a canvas object and a vertical scrollbar for scrolling it
	vscrollbar = Scrollbar(frame_data, orient=VERTICAL)
	vscrollbar.pack(fill=Y, side=RIGHT, expand=FALSE)
	
	canvas = Canvas(frame_data, bd=0, highlightthickness=0)
	canvas.configure(yscrollcommand=vscrollbar.set)
	canvas.pack()
	
	vscrollbar.config(command=canvas.yview)
	
	interieur=Frame(canvas,bg='white')
	canvas.create_window((0,0), window=interieur,anchor=NW,width=w-150,height=h-100)
	interieur.bind("<Configure>",myfunction)

	titre=Label(interieur, text="IMPORTATION DE DONNEES",bg='white',font="times 14 bold", fg="blue", width=w-250)
	titre.pack(pady=15, padx=50, side=TOP)

	panel_data=PanedWindow(interieur,bg='white', orient=HORIZONTAL, width=w-160,height=50)
	panel_data.pack(side=TOP)

	panel_tableau=PanedWindow(interieur,bg='white', orient=HORIZONTAL, width=w-160,height=50)#, bg="#aaf000"
	panel_tableau.pack(side=TOP, padx=40)

	panel_visualisation=PanedWindow(interieur,bg='white', orient=HORIZONTAL, width=w-160,height=400) #, bg="#aaf033"
	panel_visualisation.pack(side=TOP, padx=40)

	panel_entete=PanedWindow(interieur,bg='white', orient=HORIZONTAL, width=w-160,height=500) #, bg="#aaf033"
	panel_entete.pack(side=TOP, padx=40)

	#contenu de panel data
	lab_importer = LabelFrame(panel_data,bg='white', text="Choisir un fichier",width=(w-160)/4, padx=20, pady=20,relief=RIDGE)
	lab_importer.pack(expand="yes", padx=0, side=LEFT, pady=20)

	nomFichier=StringVar()
	nomFichier.set(name_fic)
	Entry(lab_importer, textvariable=nomFichier, width=28).pack(side=LEFT, padx=10)
	Button(lab_importer, text="Importer",width=10, bg='green', fg='white').pack()

	lab_colonne = LabelFrame(panel_data,bg='white', text="Nombre de colonne",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	lab_colonne.pack(expand="yes",padx=20, side=LEFT)

	nb_col=StringVar()
	nb_col.set('{} colonnes'.format(col_lig[1]))
	Label(lab_colonne, textvariable=nb_col, width=20, bg='white',fg="blue",font="times 15 bold").pack(side=LEFT)
	#Label(lab_colonne, text=" colonnes", width=7).pack(side=LEFT)

	lab_ligne = LabelFrame(panel_data,bg='white', text="Nombre de ligne",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	lab_ligne.pack(expand="yes",padx=0, side=LEFT)

	nb_lig=StringVar()
	nb_lig.set('{} lignes'.format(col_lig[0]))
	Label(lab_ligne, textvariable=nb_lig, width=20,bg='white',fg="blue",font="times 15 bold").pack(side=LEFT)
	#Label(lab_ligne, text=" lignes",width=5).pack(side=LEFT)


	Label(panel_tableau, text="CONFIGURATIONS DE LA MODELISATION",bg='white',font="times 14 bold", fg="blue", width=w-160).pack(padx=10, pady=10)


	#contenu de panel data_configuration de base
	lab_dataset = LabelFrame(panel_visualisation, bg='white',text="Type de données",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	lab_dataset.pack(expand="yes", padx=10, side=LEFT, pady=20)

	global liste_dataset

	liste_dataset=("Tabulaire","Série Temporelle")
	cb_dataset = ttk.Combobox(lab_dataset, values=liste_dataset, width=30, height=5)
	cb_dataset.set("Sélectionner")
	cb_dataset.pack()
	cb_dataset.bind('<<ComboboxSelected>>', select_dataset)


	lab_entrainement = LabelFrame(panel_visualisation,bg='white', text="Type d'entrainement",width=(w-160)/4, padx=20, pady=20,relief=RIDGE)
	lab_entrainement.pack(expand="yes",padx=10, side=LEFT, pady=20)

	global liste_entrainement
	
	liste_entrainement=("Régression","Classification")
	cb_entrainement = ttk.Combobox(lab_entrainement, values=liste_entrainement, width=30, height=5)
	cb_entrainement.set("Sélectionner")
	cb_entrainement.pack()
	cb_entrainement.bind('<<ComboboxSelected>>', select_entrainement)

	#liste_colonne=[]

	lab_index = LabelFrame(panel_visualisation,bg='white', text="Variable d'indexation",width=(w-160)/4, padx=20, pady=20,relief=RIDGE)
	lab_index.pack(expand="yes",padx=10, side=LEFT, pady=20)
	
	cb_index = ttk.Combobox(lab_index, values=liste_colonne, width=30, height=5)
	cb_index.set("Sélectionner")
	cb_index.pack()
	cb_index.bind('<<ComboboxSelected>>', select_index)

	lab_target = LabelFrame(panel_visualisation,bg='white', text="Variable cible (target)",width=(w-160)/4, padx=20, pady=20,relief=RIDGE)
	lab_target.pack(expand="yes",padx=10, side=LEFT, pady=20)

	
	cb_target = ttk.Combobox(lab_target, values=liste_colonne, width=30, height=5)
	cb_target.set("Sélectionner")
	cb_target.pack()
	cb_target.bind('<<ComboboxSelected>>', select_target)


	Label(panel_entete, text="APERCU DE DONNEES",bg='white',font="times 14 bold", fg="blue", width=w-160).pack(padx=10, pady=10)

	#afficage de fichier dans interface
	tableau = Text(panel_entete,width=w-200, height=13, fg='blue')
	tableau.pack(padx = 10, pady = 20)
	tableau.insert('end', str(data) + '\n')

	global f_progress

	f_progress=Frame(interieur,bg='white')
	f_progress.pack()

	global progress, brt

	tuple_1 = tuple(range(1, 110))
	brt = len(tuple_1)

	progress = ttk.Progressbar(f_progress, orient = HORIZONTAL,length = w-250, mode = 'determinate') 
	progress.pack(pady = 0)


	######################################################################################
	###################   FIN INTERFACE  IMPORTATION DONNEE  #############################
	######################################################################################

def general():
	############################################################################
	###################   DEBUT INTERFACE  MENU GENERAL  #######################
	############################################################################

	global pred_moyenne, dict_model, pred
	import operator
	pred_moyenne=stq.mean(liste_pred)
	pred=pred_moyenne.round(4)

	dict_model={nom:valeur for nom,valeur in zip(nom_model,liste_pred)}

	dic = OrderedDict(sorted(dict_model.items(), key=lambda x: x[1],reverse=True))

	courses = list(dic.keys()) 
	values = list(dic.values())

	# fonction qui affiche le model

	def affichage_model():
		fig_arbre=Figure(figsize=(12,5),dpi=100)

		im_arbre = fig_arbre.add_subplot(111)
		im_arbre.bar(courses, values,color='green', width = 0.4, linewidth = 5,ecolor = 'red', capsize = 30) 
		im_arbre.set_title("Vue d'ensemble des modèles")
		im_arbre.set_xlabel("Algorithmes")
		im_arbre.set_ylabel("Prédiction")
		graph_arbre = FigureCanvasTkAgg(fig_arbre, master=panel_arbre)
		canvas_arbre = graph_arbre.get_tk_widget()
		canvas_arbre.pack(padx=0)
		#aaa=fig_arbre.save(test11.PNG)


	train_sizes, train_score, val_score=learning_curve(RNA, X_train,y_train, cv=4, scoring=type_score,train_sizes=np.linspace(0.1, 1, 16))

	def courbe_d_apprentissage():
		fig_arbre=Figure(figsize=(12,5),dpi=100)

		im_arbre = fig_arbre.add_subplot(111)
		im_arbre.plot(train_sizes, train_score.mean(axis=1), color ='b',label="Score d'entrainement") # marron
		im_arbre.plot(train_sizes, val_score.mean(axis=1), color ='r',label="Score de validation")
		im_arbre.legend()
		im_arbre.set_title("Courbe d'apprentissage")
		im_arbre.set_xlabel("Taille de données entrainées")
		im_arbre.set_ylabel("Performance")
		graph_arbre = FigureCanvasTkAgg(fig_arbre, master=panel_learning_curve)
		canvas_arbre = graph_arbre.get_tk_widget()
		canvas_arbre.pack(padx=0)

	

	def myfunction(event):
		canvas.configure(scrollregion=canvas.bbox("all"),width=w-150,height=h-100)

	frame_accueil.pack_forget()
	frame_data.pack_forget()
	frame_model.pack_forget()
	frame_prevision.pack_forget()
	frame_features.pack_forget()
	frame_score.pack_forget()
	frame_historique.pack_forget()
	mb.configure(state='normal')
	
	frame_general.pack(side=TOP, padx=0)


	# create a canvas object and a vertical scrollbar for scrolling it
	vscrollbar = Scrollbar(frame_general, orient=VERTICAL)
	vscrollbar.pack(fill=Y, side=RIGHT, expand=FALSE)

	canvas = Canvas(frame_general, bd=0, highlightthickness=0)
	canvas.configure(yscrollcommand=vscrollbar.set)
	canvas.pack()

	vscrollbar.config(command=canvas.yview)

	interieur=Frame(canvas, bg='white')
	canvas.create_window((0,0), window=interieur,anchor=NW,width=w-150,height=h+600)
	interieur.bind("<Configure>",myfunction)


	titre=Label(interieur, text="REPRESENTATION DES MODELES",bg='white', fg="blue",font="times 15 bold")
	titre.pack(pady=20, side=TOP)

	panel_train_score=PanedWindow(interieur, bg='white',orient=HORIZONTAL, width=w-160,height=50)
	panel_train_score.pack(side=TOP)

	panel_arbre=PanedWindow(interieur,bg='white', orient=HORIZONTAL, width=w-200,height=200) #, bg="#aaf033"
	panel_arbre.pack(side=TOP, padx=40, pady=20)

	affichage_model()

	panel_titre=PanedWindow(interieur, bg='white',orient=HORIZONTAL, width=w-200,height=200) #, bg="#aaf033"
	panel_titre.pack(side=TOP, padx=40, pady=10)

	panel_learning_curve=PanedWindow(interieur,bg='white', orient=HORIZONTAL, width=w-200,height=400)#, bg="#aaf000"
	panel_learning_curve.pack(side=TOP, padx=20)

	courbe_d_apprentissage()

	lab_score = LabelFrame(panel_train_score,bg='white', text="Estimation moyenne",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	lab_score.pack(expand="yes",padx=20, side=LEFT)

	score_estime=StringVar()
	
	if type_dataset==liste_dataset[0]:
		if type_entrainement==liste_entrainement[0]:
			estimation=pred
			score_estime.set(estimation)
		else:
			if pred<=0.5:
				estimation="Echec"
				score_estime.set("{} => {}".format(pred,estimation))
			else:
				estimation="Succès"
				score_estime.set("{} => {}".format(pred,estimation))
	else:
		estimation=pred
		score_estime.set(estimation)

	
	
	#score_estime.set(pred_tree)
	Label(lab_score, textvariable=score_estime, width=20,bg='white', fg="blue",font="times 15 bold").pack(side=LEFT)
	#Label(lab_score, text=" colonnes", width=n7).pack(side=LEFT)

	lab_nomtarget = LabelFrame(panel_train_score,bg='white', text="Variable selectionnée à predire",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	lab_nomtarget.pack(expand="yes",padx=20, side=LEFT)
	
	nom_target=StringVar()
	nom_target.set(type_target)
	Label(lab_nomtarget, textvariable=nom_target,bg='white', width=20, fg="blue",font="times 15 bold").pack(side=LEFT)

	lab_model = LabelFrame(panel_train_score,bg='white', text="Nombre de modeles entrainés",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	lab_model.pack(expand="yes",padx=0, side=LEFT)

	global nb_model
	nb_model=len(nom_model)
	nb_model_est=StringVar()
	nb_model_est.set(nb_model)
	Label(lab_model, textvariable=nb_model_est,bg='white', width=20, fg="blue",font="times 15 bold").pack(side=LEFT)
	#Label(lab_model, text=" lignes",width=5).pack(side=LEFT)


	titre_tab=StringVar()
	titre_tab.set("APPRENTISSAGE")

	Label(panel_titre, textvariable=titre_tab,font="times 14 bold",bg='white', fg="blue", width=w-160).pack(padx=10, pady=20)


	############################################################################
	#####################   FIN INTERFACE  MENU GENERAL ########################
	############################################################################

def model():

	def hyperparametres(typ=None,tech=None,mod=None, scr=None,mtrq_dev=None, t=None,deploy=None,statu=None,
		r_state=None,kern=None,loss_mod=None, mx_iter=None, normalise_y=None,tol_mod=None,neighbors=None,weights_mod=None,
		leaf=None, algo=None,mx_feature=None, mx_depth=None, min_samples=None, solver_mod=None,hidden_size=None,
		estimateur=None,learn_rate=None, nstart=None, a=None):
		type_modele_sélectionner.set(typ)
		technologie.set(tech)
		module.set(mod)
		score_model.set(scr)
		metrique_model.set(mtrq)
		duree_train.set(t)
		deployement.set(deploy)
		statut.set(statu)
		Polynomial.set("2")
		selectKbest.set("6")
		random_stt.set(r_state)
		kernel.set(kern)
		loss.set(loss_mod)
		max_iter.set(mx_iter)
		normalize_y.set(normalise_y)
		tol.set(tol_mod)
		n_neighbors.set(neighbors)
		weights.set(weights_mod)
		leaf_size.set(leaf)
		algorithm.set(algo)
		max_features.set(mx_feature)
		max_depth.set(mx_depth)
		min_samples_split.set(min_samples)
		solver.set(solver_mod)
		hidden_layer_sizes.set(hidden_size)
		n_estimators.set(estimateur)
		learning_rate.set(learn_rate)
		n_restarts_optimizer.set(nstart)
		alpha.set(a)

	#selection de modele entrainner
	def entrain_model(event=None):
		global model_selectionne
		if event:
			model_selectionne=event.widget.get()

			if model_selectionne=='Ridge':
				hyperparametres(typ='Lineaire',tech='Scikit-learn',mod='linear_model', scr=score_ridge,mtrq_dev=None, t=t_ridge,deploy='Oui',statu='Actif',
				r_state=0,kern=None,loss_mod=None, mx_iter=None, normalise_y=None,tol_mod=None,neighbors=None,weights_mod=None,
				leaf=None, algo=None,mx_feature=None, mx_depth=None, min_samples=None, solver_mod=None,hidden_size=None,
				estimateur=None,learn_rate=None, nstart=None, a=None)
			if model_selectionne=='SGD':
				hyperparametres(typ='Lineaire',tech='Scikit-learn',mod='linear_model', scr=score_sgd,mtrq_dev=None, t=t_sgd,deploy='Oui',statu='Actif',
				r_state=0,kern=None,loss_mod="squard_loss", mx_iter=1000, normalise_y=None,tol_mod=1e-3,neighbors=None,weights_mod=None,
				leaf=None, algo=None,mx_feature=None, mx_depth=None, min_samples=None, solver_mod=None,hidden_size=None,
				estimateur=None,learn_rate=None, nstart=None, a=None)
			if model_selectionne=='SVM':
				hyperparametres(typ='SVM',tech='Scikit-learn',mod='SVM', scr=score_svm,mtrq_dev=None, t=t_svm,deploy='Oui',statu='Actif',
				r_state=None,kern='linear',loss_mod=None, mx_iter=None, normalise_y=None,tol_mod=None,neighbors=None,weights_mod=None,
				leaf=None, algo=None,mx_feature=None, mx_depth=None, min_samples=None, solver_mod=None,hidden_size=None,
				estimateur=None,learn_rate=None, nstart=None, a=None)
			if model_selectionne=='KNN':
				hyperparametres(typ='Neighbors',tech='Scikit-learn',mod='neighbors', scr=score_knn,mtrq_dev=None, t=t_knn,deploy='Oui',statu='Actif',
				r_state=0,kern=None,loss_mod=None, mx_iter=None, normalise_y=None,tol_mod=None,neighbors=2,weights_mod='uniform',
				leaf=30, algo='auto',mx_feature=None, mx_depth=None, min_samples=None, solver_mod=None,hidden_size=None,
				estimateur=None,learn_rate=None, nstart=None, a=None)
			if model_selectionne=='Tree':
				hyperparametres(typ='Arbre',tech='Scikit-learn',mod='tree', scr=score_tree,mtrq_dev=None, t=t_tree,deploy='Oui',statu='Actif',
				r_state=0,kern=None,loss_mod=None, mx_iter=1000, normalise_y=None,tol_mod=None,neighbors=None,weights_mod=None,
				leaf=None, algo=None,mx_feature=5, mx_depth=4, min_samples=5, solver_mod=None,hidden_size=None,
				estimateur=None,learn_rate=None, nstart=None, a=None)
			if model_selectionne=='Extratree':
				hyperparametres(typ='Arbre',tech='Scikit-learn',mod='ensemble', scr=score_extratree,mtrq_dev=None, t=t_extratree,deploy='Oui',statu='Actif',
				r_state=0,kern=None,loss_mod=None, mx_iter=None, normalise_y=None,tol_mod=None,neighbors=None,weights_mod=None,
				leaf=None, algo=None,mx_feature=5, mx_depth=4, min_samples=5, solver_mod=None,hidden_size=None,
				estimateur=None,learn_rate=None, nstart=None, a=None)
			if model_selectionne=='RNA':
				hyperparametres(typ='Reseau de Neurone',tech='Scikit-learn',mod='neural_network', scr=score_rna,mtrq_dev=None, t=t_rna,deploy='Oui',statu='Actif',
				r_state=0,kern=None,loss_mod=None, mx_iter=1000, normalise_y=None,tol_mod=None,neighbors=None,weights_mod=None,
				leaf=None, algo=None,mx_feature=None, mx_depth=None, min_samples=None, solver_mod='lbfgs',hidden_size='(8,)',
				estimateur=None,learn_rate=None, nstart=None, a=None)
			if model_selectionne=='RForest':
				hyperparametres(typ='Foret aleatoire',tech='Scikit-learn',mod='ensemble', scr=score_rforest,mtrq_dev=None, t=t_rforest,deploy='Oui',statu='Actif',
				r_state=0,kern=None,loss_mod=None, mx_iter=None, normalise_y=None,tol_mod=None,neighbors=None,weights_mod=None,
				leaf=None, algo=None,mx_feature=None, mx_depth=None, min_samples=None, solver_mod=None,hidden_size=None,
				estimateur=None,learn_rate=None, nstart=None, a=None)
			if model_selectionne=='GBoosting':
				hyperparametres(typ='Boosting',tech='Scikit-learn',mod='ensemble', scr=score_gboosting,mtrq_dev=None, t=t_gboosting,deploy='Oui',statu='Actif',
				r_state=0,kern=None,loss_mod=None, mx_iter=None, normalise_y=None,tol_mod=None,neighbors=None,weights_mod=None,
				leaf=None, algo=None,mx_feature=None, mx_depth=None, min_samples=None, solver_mod=None,hidden_size=None,
				estimateur=None,learn_rate=None, nstart=None, a=None)
			if model_selectionne=='AdaBoost':
				hyperparametres(typ='Boosting',tech='Scikit-learn',mod='ensemble', scr=score_adaboost,mtrq_dev=None, t=t_adaboost,deploy='Oui',statu='Actif',
				r_state=0,kern=None,loss_mod='linear', mx_iter=None, normalise_y=None,tol_mod=None,neighbors=None,weights_mod=None,
				leaf=None, algo=None,mx_feature=None, mx_depth=None, min_samples=None, solver_mod=None,hidden_size=None,
				estimateur=40,learn_rate=1.0, nstart=None, a=None)
			if model_selectionne=='Gaussian':
				hyperparametres(typ='Gaussian',tech='Scikit-learn',mod='gaussian_process', scr=score_gaussian,mtrq_dev=None, t=t_gaussian,deploy='Oui',statu='Actif',
				r_state=0,kern='DotProduct() + WhiteKernel()',loss_mod=None, mx_iter=None, normalise_y='True',tol_mod=None,neighbors=None,weights_mod=None,
				leaf=None, algo=None,mx_feature=None, mx_depth=None, min_samples=None, solver_mod=None,hidden_size=None,
				estimateur=None,learn_rate=None, nstart=0, a=0.5)
			if model_selectionne=='NBayes':
				hyperparametres(typ='Bayesien',tech='Scikit-learn',mod='naive_bayes', scr=score_bayes,mtrq_dev=None, t=t_bayes,deploy='Oui',statu='Actif',
				r_state=None,kern=None,loss_mod=None, mx_iter=None, normalise_y=None,tol_mod=None,neighbors=None,weights_mod=None,
				leaf=None, algo=None,mx_feature=None, mx_depth=None, min_samples=None, solver_mod=None,hidden_size=None,
				estimateur=None,learn_rate=None, nstart=None, a=None)

	

	##############################################################################
	##############  DEBUT INTERFACE DE MODEL     #################################
	##############################################################################
	def myfunction(event):
		canvas.configure(scrollregion=canvas.bbox("all"),width=w-150,height=h-100)

	frame_general.pack_forget()
	mb.configure(state='normal')
	frame_accueil.pack_forget()
	frame_data.pack_forget()
	frame_prevision.pack_forget()
	frame_features.pack_forget()
	frame_score.pack_forget()
	frame_historique.pack_forget()

	frame_model.pack(side=TOP, padx=0)


	# create a canvas object and a vertical scrollbar for scrolling it
	vscrollbar = Scrollbar(frame_model, orient=VERTICAL)
	vscrollbar.pack(fill=Y, side=RIGHT, expand=FALSE)

	canvas = Canvas(frame_model, bd=0, highlightthickness=0)
	canvas.configure(yscrollcommand=vscrollbar.set)
	canvas.pack()

	vscrollbar.config(command=canvas.yview)

	intern_frame=Frame(canvas)
	canvas.create_window((0,0), window=intern_frame,anchor=NW,width=w-150,height=h-100)
	intern_frame.bind("<Configure>",myfunction)


	titre=Label(intern_frame, text="CARACTERISTIQUES DES MODELES", fg="blue",font="times 15 bold")
	titre.pack(pady=10, side=TOP)

	panel_tab_model=PanedWindow(intern_frame, orient=HORIZONTAL, width=w-160,height=800) #, bg="#aaf033"
	panel_tab_model.pack(side=TOP, padx=40, pady=0)

	panel_description=PanedWindow(intern_frame, orient=HORIZONTAL, width=w-160,height=200)#, bg="#aaf000"
	panel_description.pack(side=TOP, padx=40)

	panel_info_hyperparam=PanedWindow(intern_frame, orient=HORIZONTAL, width=w-160,height=400)#, bg="#aaf000"
	panel_info_hyperparam.pack(side=TOP, padx=40)

	
	#tableau = ttk.Treeview(panel_tab_model, columns=('nom', 'tech', 'type', 'score','duree', 'statut'))

	tableau = ttk.Treeview(panel_tab_model,selectmode="browse", height=12)

	tableau["columns"]=("nom", "tech", "type","prediction",
				"score", "duree", "statut")
	tableau['show'] = 'headings'

	tableau.column('nom', width=140)
	tableau.heading('nom', text='Nom du modèle')

	tableau.column('tech', width=140,anchor="center")
	tableau.heading('tech', text='Téchnologie')

	tableau.column('type', width=140 ,anchor="center")
	tableau.heading('type', text='Type')

	tableau.column('prediction', width=140, anchor="center")
	tableau.heading('prediction', text='Valeur prédite')

	tableau.column('score', width=140, anchor="center")
	tableau.heading('score', text='Score')

	tableau.column('duree', width=140, anchor="center")
	tableau.heading('duree', text='Durée de prédiction (s)')

	tableau.column('statut', width=140, anchor="center")
	tableau.heading('statut', text='Statut')

	#tableau.bind("<Double-1>", clic_tableau)
	#tableau.grid(row=1, rowspan=3, column=1)
	tableau.pack(padx = 10, pady = (0, 10))

	tab_ridge=["Ridge","SKLEARN","Linéaire",pred_ridge,score_ridge,t_ridge,"Actif"]
	tab_sgd=["SGD","SKLEARN","Linéaire",pred_sgd,score_sgd,t_sgd,"Actif"]
	tab_svm=["SVM","SKLEARN","SVM",pred_svm,score_svm,t_svm,"Actif"]
	tab_knn=["KNN","SKLEARN","Neighbors",pred_knn,score_knn,t_knn,"Actif"]

	tab_tree=["Tree","SKLEARN","Arbre",pred_tree,score_tree,t_tree,"Actif"]
	tab_extratree=["Extratree","SKLEARN","Arbre",pred_extratree,score_extratree,t_extratree,"Actif"]
	tab_rna=["RNA","SKLEARN","Réseau de Neurone",pred_rna,score_rna,t_rna,"Actif"]
	tab_rforest=["RForest","SKLEARN","Foret Aléatoire",pred_rforest,score_rforest,t_rforest,"Actif"]
	tab_gboosting=["GBoosting","SKLEARN","Boosting",pred_gboosting,score_gboosting,t_gboosting,"Actif"]
	tab_adaboost=["AdaBoost","SKLEARN","Boosting",pred_adaboost,score_adaboost,t_adaboost,"Actif"]
	tab_gaussian=["Gaussian","SKLEARN","Gaussian",pred_gaussian,score_gaussian,t_gaussian,"Actif"]

	global pred_bayes, statut_bayes, score_bayes, t_bayes
	if type_dataset==liste_dataset[0]:
		if type_entrainement==liste_entrainement[0]:
			pred_bayes='Indéfinie'
			statut_bayes='Inactif'
			score_bayes='Indéfinie'
			t_bayes='Indéfinie'
		else:
			statut_bayes='Actif'
	if type_dataset==liste_dataset[1]:
		pred_bayes='Indéfinie'
		statut_bayes='Inactif'
		score_bayes='Indéfinie'
		t_bayes='Indéfinie'

	tab_nbayes=["NBayes","SKLEARN","Bayesian",pred_bayes,score_bayes,t_bayes,statut_bayes]

	tableau.insert("","end", values=tab_ridge)
	tableau.insert("","end", values=tab_sgd)
	tableau.insert("","end", values=tab_svm)
	tableau.insert("","end", values=tab_knn)
	tableau.insert("","end", values=tab_tree)
	tableau.insert("","end", values=tab_extratree)
	tableau.insert("","end", values=tab_rna)
	tableau.insert("","end", values=tab_rforest)
	tableau.insert("","end", values=tab_gboosting)
	tableau.insert("","end", values=tab_adaboost)
	tableau.insert("","end", values=tab_gaussian)
	tableau.insert("","end",values=tab_nbayes)


	#titre_tab=StringVar()
	# INFORMATION DU MODEL

	Label(panel_description, text="INFORMATION ET LES HYPERPARAMETRES DU MODELE ",font="times 14 bold", fg="blue", width=w-250).pack(padx=10, pady=20)

	info_frame=LabelFrame(panel_info_hyperparam, text="Information du modèle",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	info_frame.pack(expand="yes",padx=20, side=LEFT)


	#Nom de model
	a=Frame(info_frame)
	a.pack(side=TOP)
	Label(a,text="Nom : ").pack(side=LEFT)

	#_model=("Regression","Classification")
	cb = ttk.Combobox(a, values=nom_model, width=30, height=5)
	cb.set("Sélectionner")
	cb.pack()
	cb.bind('<<ComboboxSelected>>', entrain_model)


	#TYPE DU MODEL
	a=Frame(info_frame)
	a.pack(side=TOP)
	Label(a,text="Type du modèle :   ").pack(side=LEFT)

	type_modele_sélectionner=StringVar()
	Label(a, textvariable=type_modele_sélectionner).pack(side=LEFT)

	#Technologie
	a=Frame(info_frame)
	a.pack(side=TOP)
	Label(a,text="Téchnologie :     ").pack(side=LEFT)

	technologie=StringVar()
	Label(a, textvariable=technologie).pack(side=LEFT)

	#Module
	a=Frame(info_frame)
	a.pack(side=TOP)
	Label(a,text="Module :").pack(side=LEFT)

	module=StringVar()
	#module.set("ENSEMBLE")
	Label(a, textvariable=module).pack(side=LEFT)

	#SCORE
	b=Frame(info_frame)
	b.pack(side=TOP)
	Label(b,text="Score :          ").pack(side=LEFT)

	score_model=StringVar()
	Label(b, textvariable=score_model).pack(side=LEFT)

	#metrique
	c=Frame(info_frame)
	c.pack(side=TOP)
	Label(c,text="Metrique :          ").pack(side=LEFT)

	metrique_model=StringVar()
	#metrique_model.set("ENSEMBLE")
	Label(c, textvariable=metrique_model).pack(side=LEFT)

	#Randome_state
	c1=Frame(info_frame)
	c1.pack(side=TOP)
	Label(c1,text="random_state :          ").pack(side=LEFT)

	random_stt=StringVar()
	Label(c1, textvariable=random_stt).pack(side=LEFT)


	#Duree d'entrainement
	e=Frame(info_frame)
	e.pack(side=TOP)
	Label(e,text="Durée d'entrainement:").pack(side=LEFT)

	duree_train=StringVar()
	#duree_train.set("ENSEMBLE")
	Label(e, textvariable=duree_train).pack(side=LEFT)


	#Deployement
	e=Frame(info_frame)
	e.pack(side=TOP)
	Label(e,text="Deployable:").pack(side=LEFT)

	deployement=StringVar()
	Label(e, textvariable=deployement).pack(side=LEFT)

	#Statut
	f=Frame(info_frame)
	f.pack(side=TOP)
	Label(f,text="Statut :                 ").pack(side=LEFT)

	statut=StringVar()
	Label(f, textvariable=statut).pack(side=LEFT)

	################# Hyperparametres   ###################

	hyp_frame=LabelFrame(panel_info_hyperparam, text="Hypérparametres du modèle",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	hyp_frame.pack(expand="yes",padx=50, side=LEFT)

	frame1=Frame(hyp_frame)
	frame1.pack(side=LEFT, padx=20)

	#Polynomial
	a1=Frame(frame1)
	a1.pack(side=TOP)
	Label(a1,text="Polynomial features :  ").pack(side=LEFT)

	Polynomial=StringVar()
	Label(a1, textvariable=Polynomial).pack(side=LEFT)

	#selectKbest
	b1=Frame(frame1)
	b1.pack(side=TOP)
	Label(b1,text="SelectKbest  :          ").pack(side=LEFT)

	selectKbest=StringVar()
	Label(b1, textvariable=selectKbest).pack(side=LEFT)


	#kernel
	d1=Frame(frame1)
	d1.pack(side=TOP)
	Label(d1,text="Kernel :        ").pack(side=LEFT)

	kernel=StringVar()
	Label(d1, textvariable=kernel).pack(side=LEFT)

	#Loss
	e1=Frame(frame1)
	e1.pack(side=TOP)
	Label(e1,text="loss :").pack(side=LEFT)

	loss=StringVar()
	Label(e1, textvariable=loss).pack(side=LEFT)

	#Max_iter
	f1=Frame(frame1)
	f1.pack(side=TOP)
	Label(f1,text="max_iter :                 ").pack(side=LEFT)

	max_iter=StringVar()
	Label(f1, textvariable=max_iter).pack(side=LEFT)

	#normalize_y
	f1=Frame(frame1)
	f1.pack(side=TOP)
	Label(f1,text="normalize_y :                 ").pack(side=LEFT)

	normalize_y=StringVar()
	Label(f1, textvariable=normalize_y).pack(side=LEFT)

	#Tol
	f1=Frame(frame1)
	f1.pack(side=TOP)
	Label(f1,text="Tol :                 ").pack(side=LEFT)

	tol=StringVar()
	Label(f1, textvariable=tol).pack(side=LEFT)


	#n_neighbors
	f1=Frame(frame1)
	f1.pack(side=TOP, padx=30)
	Label(f1,text="n_neighbors :                 ").pack(side=LEFT)

	n_neighbors=StringVar()
	Label(f1, textvariable=n_neighbors).pack(side=LEFT)

	#weights
	f1=Frame(frame1)
	f1.pack(side=TOP)
	Label(f1,text="weights :                 ").pack(side=LEFT)

	weights=StringVar()
	Label(f1, textvariable=weights).pack(side=LEFT)

	#leaf_size
	f1=Frame(frame1)
	f1.pack(side=TOP)
	Label(f1,text="leaf_size :                 ").pack(side=LEFT)

	leaf_size=StringVar()
	Label(f1, textvariable=leaf_size).pack(side=LEFT)

	####### 2e Frame d'hyperparam  ###########
	frame2=Frame(hyp_frame)
	frame2.pack(side=LEFT, padx=30)

	#algorithm
	f1=Frame(frame2)
	f1.pack(side=TOP)
	Label(f1,text="algorithm :                 ").pack(side=LEFT)

	algorithm=StringVar()
	Label(f1, textvariable=algorithm).pack(side=LEFT)


	#max_features
	f1=Frame(frame2)
	f1.pack(side=TOP)
	Label(f1,text="max_features :                 ").pack(side=LEFT)

	max_features=StringVar()
	Label(f1, textvariable=max_features).pack(side=LEFT)


	#max_depth
	f1=Frame(frame2)
	f1.pack(side=TOP)
	Label(f1,text="max_depth :                 ").pack(side=LEFT)

	max_depth=StringVar()
	Label(f1, textvariable=max_depth).pack(side=LEFT)


	#min_samples_split
	f1=Frame(frame2)
	f1.pack(side=TOP)
	Label(f1,text="min_samples_split :                 ").pack(side=LEFT)

	min_samples_split=StringVar()
	Label(f1, textvariable=min_samples_split).pack(side=LEFT)


	#solver
	f1=Frame(frame2)
	f1.pack(side=TOP)
	Label(f1,text="solver  :                 ").pack(side=LEFT)

	solver=StringVar()
	Label(f1, textvariable=solver).pack(side=LEFT)


	#hidden_layer_sizes
	f1=Frame(frame2)
	f1.pack(side=TOP)
	Label(f1,text="hidden_layer_sizes :                 ").pack(side=LEFT)

	hidden_layer_sizes=StringVar()
	Label(f1, textvariable=hidden_layer_sizes).pack(side=LEFT)

	#n_estimators
	f1=Frame(frame2)
	f1.pack(side=TOP)
	Label(f1,text="n_estimators :                 ").pack(side=LEFT)

	n_estimators=StringVar()
	Label(f1, textvariable=n_estimators).pack(side=LEFT)

	#learning_rate
	f1=Frame(frame2)
	f1.pack(side=TOP)
	Label(f1,text="learning_rate :                 ").pack(side=LEFT)

	learning_rate=StringVar()
	Label(f1, textvariable=learning_rate).pack(side=LEFT)

	#n_restarts_optimizer
	f1=Frame(frame2)
	f1.pack(side=TOP)
	Label(f1,text="n_restarts_optimizer :                 ").pack(side=LEFT)

	n_restarts_optimizer=StringVar()
	Label(f1, textvariable=n_restarts_optimizer).pack(side=LEFT)

	#alpha
	f1=Frame(frame2)
	f1.pack(side=TOP)
	Label(f1,text="alpha :                 ").pack(side=LEFT)

	alpha=StringVar()
	Label(f1, textvariable=alpha).pack(side=LEFT)

	

	##############################################################################
	##############  FIN INTERFACE DE MODEL     #################################
	##############################################################################

def prevision():
	############################################################################
	###################   DEBUT INTERFACE  MENU PREVISION  #######################
	############################################################################

	def selection_model(event=None):
		global date_prev
		if event:
			date_prev=event.widget.get()


	def myfunction(event):
		canvas.configure(scrollregion=canvas.bbox("all"),width=w-150,height=h-100)

	#frame_prevision=Frame(fenetre)
	frame_prevision.pack(side=TOP, padx=0)
	mb.configure(state='normal')
	frame_data.pack_forget()
	frame_model.pack_forget()
	frame_general.pack_forget()
	frame_accueil.pack_forget()
	frame_features.pack_forget()
	frame_score.pack_forget()
	frame_historique.pack_forget()


	# create a canvas object and a vertical scrollbar for scrolling it
	vscrollbar = Scrollbar(frame_prevision, orient=VERTICAL)
	vscrollbar.pack(fill=Y, side=RIGHT, expand=FALSE)

	canvas = Canvas(frame_prevision, bd=0, highlightthickness=0)
	canvas.configure(yscrollcommand=vscrollbar.set)
	canvas.pack()

	vscrollbar.config(command=canvas.yview)

	interieur=Frame(canvas,bg='white')
	canvas.create_window((0,0), window=interieur,anchor=NW,width=w-150,height=h*2+150)
	interieur.bind("<Configure>",myfunction)


	titre=Label(interieur, text="PAGE DE LA PREVISION", bg='white',fg="blue",font="times 15 bold")
	titre.pack(pady=25, side=TOP)

	panel_selection_date=PanedWindow(interieur, orient=HORIZONTAL,bg='white', width=w-160,height=50)
	panel_selection_date.pack(side=TOP)

	panel_graphe=PanedWindow(interieur, orient=HORIZONTAL,bg='white', width=w-160,height=400) #, bg="#aaf033"
	panel_graphe.pack(side=TOP, padx=40, pady=20)


	lab_select_model = LabelFrame(panel_selection_date,bg='white', text="Paramètres de prévision",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	lab_select_model.pack(expand="yes",padx=10, side=LEFT)

	list_jour=['Index','Jour','Semaine','Mois','Trimestre','Semestre', 'Année']

	def select_unite(event):
		global unite, jour_pred
		if event:
			unite=event.widget.get()

			if unite==list_jour[0]: #semaine
				jour_pred=7
			if unite==list_jour[1]: #mois
				jour_pred=30
			if unite==list_jour[2]: #trimestre
				jour_pred=90
			if unite==list_jour[3]: #semestre
				jour_pred=180
			if unite==list_jour[4]: #anne
				jour_pred=356
			if unite==list_jour[5]: #anne
				jour_pred=712

	def select_model_pref(event):
		global mod_pref
		if event:
			mod_pref=event.widget.get()
			

	if type_dataset==liste_dataset[0]:
		list_jour=['index']
		val_mod_pred=list(dict_model.keys())


	if type_dataset==liste_dataset[1]:
		list_jour=['1 Semaine','1 Mois','1 Trimestre','1 Semestre', '1 An', '2 Ans']
		val_mod_pred=list(dict_model.keys())


	def select_x(event):
		global var_x
		if event:
			var_x=event.widget.get()


	combo_unite = ttk.Combobox(lab_select_model, values=list_jour, width=10, height=5)
	combo_unite.set("Unité")
	combo_unite.bind('<<ComboboxSelected>>', select_unite)

	combo_cfr = ttk.Combobox(lab_select_model, values=val_mod_pred, width=10, height=5)
	combo_cfr.set("Modele")
	combo_cfr.pack(padx=10, side=LEFT)
	combo_cfr.bind('<<ComboboxSelected>>', select_model_pref)

	combo_x = ttk.Combobox(lab_select_model, values=liste_colonne, width=10, height=5)
	combo_x.set("Variable x")
	#combo_x.pack(padx=20, side=LEFT)
	combo_x.bind('<<ComboboxSelected>>', select_x)

	val_predire=Entry(lab_select_model, width=10)


	if type_dataset==liste_dataset[0]:
		if type_entrainement==liste_entrainement[0]:
			combo_x.pack(padx=20, side=LEFT)
			#combo_unite.pack_forget(padx=10, side=LEFT)
			val_predire.pack_forget()

		else:
			combo_x.pack(padx=20, side=LEFT)
			#combo_unite.pack_forget()
			val_predire.pack()

	else:
		combo_x.pack_forget()
		val_predire.pack_forget()
		combo_unite.pack(padx=10, side=LEFT)


	def prevision_ARIMA(prev):

		def obtenir_prevision(y,max):
			import statsmodels.api as sm
			mod = sm.tsa.statespace.SARIMAX(y,
										order=(3, 1, 1),
										seasonal_order=(2, 1, 1, 5),
										enforce_stationarity=False,
										enforce_invertibility=False)

			results = mod.fit()

			#print(results.summary().tables[1])
			pred_uc = results.get_forecast(steps=max)

			# Get confidence intervals of forecasts
			pred_ci = pred_uc.conf_int()
			prevision=pred_uc.predicted_mean

			return prevision

		pred=obtenir_prevision(y_reel,jour_pred)
		pred_concat=list(y_reel)+list(pred)

		f_prev = Figure(figsize=(12,5), dpi=100)

		pp = f_prev.add_subplot(111)

		pp.plot(pred_concat, label='Prévision' ,lw=3, c='orange')#, c='green'
		pp.plot(y_reel,label='Variable réelle ({})'.format(type_target),lw=3)
		#pp.set_yticklabels('Valeur')
		pp.set_xlabel("Période")
		pp.set_ylabel("Valeur")

		
		pp.legend(loc='best')
		pp.set_title('Graphe de prévision')

		grp = FigureCanvasTkAgg(f_prev, master=panel_graphe)
		canv3=grp.get_tk_widget().pack()


	def faire_predire():

		if type_dataset==liste_dataset[0]:

			if type_entrainement==liste_entrainement[1]:
				x_saisie=val_predire.get()
				data_x=data.drop(type_target, axis=1)
				moyenne_col=data_x.mean()

				dic_moyenne={nom:valeur for nom,valeur in zip(liste_colonne,moyenne_col)}

				key_dic_moyenne=list(dic_moyenne.keys())
				value_dic_moyenne=list(dic_moyenne.values())

				for i in range(len(dic_moyenne)):
					if var_x==key_dic_moyenne[i]:
						pp=value_dic_moyenne[i]

						dic_moyenne[var_x]=x_saisie

						x_pp=np.array(list(dic_moyenne.values())).reshape(1,len(list(dic_moyenne.values())))
						
						if mod_pref== 'Ridge':  
							pr=Ridge.predict(x_pp);
						if mod_pref== 'SGD':  
							pr=SGD.predict(x_pp);
						if mod_pref==  'SVM': 
							pr=SVM.predict(x_pp);
						if mod_pref== 'KNN':  
							pr=KNN.predict(x_pp);
						if mod_pref== 'Tree':  
							pr=Tree.predict(x_pp);
						if mod_pref== 'Extratree':  
							pr=Extratree.predict(x_pp);
						if mod_pref== 'RNA':  
							pr=RNA.predict(x_pp);
						if mod_pref== 'RForest':  
							pr=RForest.predict(x_pp);
						if mod_pref== 'GBoosting':  
							pr=GBoosting.predict(x_pp);
						if mod_pref== 'AdaBoost': 
							pr=AdaBoost.predict(x_pp);
						if mod_pref== 'Gaussian': 
							pr=Gaussian.predict(x_pp);
						if mod_pref== 'NBayes': 
							pr=Bayes.predict(x_pp);

						if pr==1:
							resultat.set("Succès")
						if pr==0:
							resultat.set("Echec")

			else:

				list_mod=list(dict_model.keys())
				val_mod=list(dict_model.values())

				for i in range(len(dict_model)):
					if mod_pref== list_mod[i]:
						y_res=val_mod[i]
						resultat.set(y_res)
		else:
			list_mod=list(dict_model.keys())
			val_mod=list(dict_model.values())

			for i in range(len(dict_model)):
				if mod_pref== list_mod[i]:
					y_res=val_mod[i]
					resultat.set(y_res)

	def prevision_classification():
		f_prev = Figure(figsize=(10,5), dpi=100)

		pp = f_prev.add_subplot(111)

		pp.plot(SMA, label='Prévision Moyenne Mobile' ,lw=3, c='orange')#, c='green'
		pp.plot(y_reel,label='Variable réelle ({})'.format(type_target))
		#pp.set_yticklabels('Valeur')
		pp.set_xlabel("Variation")
		pp.set_ylabel("Valeur")

		
		pp.legend(loc='best')
		pp.set_title('Graphe de prévision')

		grp = FigureCanvasTkAgg(f_prev, master=panel_graphe)
		canv3=grp.get_tk_widget().grid()


	def prevision_regression():
		f_prev = Figure(figsize=(10,5), dpi=100)

		pp = f_prev.add_subplot(111)

		pp.plot(SMA, label='Prévision Moyenne Mobile' ,lw=3, c='orange')#, c='green'
		pp.plot(y_reel,label='Variable réelle ({})'.format(type_target))
		#pp.set_yticklabels('Valeur')
		pp.set_xlabel("Variation")
		pp.set_ylabel("Valeur")

		
		pp.legend(loc='best')
		pp.set_title('Graphe de prévision')

		grp = FigureCanvasTkAgg(f_prev, master=panel_graphe)
		canv3=grp.get_tk_widget().grid()
		#canv3.pack(side=LEFT,padx=10)

	lab_model = LabelFrame(panel_selection_date,bg='white', text="Commande de prévision",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	lab_model.pack(expand="yes",padx=10, side=LEFT)

	Button(lab_model, text="Prédire", width=15, fg="white", bg='green', command=faire_predire).pack(side=LEFT, padx=10)
	#Label(lab_model, text=" lignes",width=5).pack(side=LEFT)



	lab_model = LabelFrame(panel_selection_date,bg='white', text="Résultat de prédiction ",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	lab_model.pack(expand="yes",padx=10, side=LEFT)

	resultat=StringVar()
	resultat.set('None')

	Label(lab_model, textvariable=resultat,width=15,bg='white', font='sans 14 bold', fg="blue").pack()

	def animation():
		if type_dataset==liste_dataset[0]:
			if type_entrainement==liste_entrainement[0]:
				prevision_regression()
			else:
				prevision_classification()
		else:
			#prevision_regression()
			prevision_ARIMA(jour_pred)
		
	lab_model_cmd = LabelFrame(panel_selection_date,bg='white', text="Représentation graphique ",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	lab_model_cmd.pack(expand="yes",padx=10, side=LEFT)

	cmd=Button(lab_model_cmd, text="Animer la graphe", width=15, fg="white", bg='green', command=animation)
	#Label(lab_model, text=" lignes",width=5).pack(side=LEFT)
	cmd.pack(side=LEFT, padx=10)

	# GRAPHE DE PREVISION
	y_reel=list(data[type_target])

	SMA = data[type_target].rolling(window=5).mean()

	#SMA=data[type_target].ewm(span=40,adjust=False).mean()

	############################################################################
	###################   FIN INTERFACE  MENU PREVISION ########################
	############################################################################

def features():
	############################################################################
	###################   DEBUT INTERFACE  MENU FEATURES  #######################
	############################################################################

	def selection_model(event=None):
		global date_prev
		if event:
			date_prev=event.widget.get()


	def myfunction(event):
		canvas.configure(scrollregion=canvas.bbox("all"),width=w-150,height=h-100)

	
	frame_features.pack(side=TOP, padx=0)
	mb.configure(state='normal')
	frame_data.pack_forget()
	frame_general.pack_forget()
	frame_model.pack_forget()
	frame_prevision.pack_forget()
	frame_accueil.pack_forget()
	frame_score.pack_forget()
	frame_historique.pack_forget()

	# create a canvas object and a vertical scrollbar for scrolling it
	vscrollbar = Scrollbar(frame_features, orient=VERTICAL)
	vscrollbar.pack(fill=Y, side=RIGHT, expand=FALSE)

	canvas = Canvas(frame_features, bd=0, highlightthickness=0)
	canvas.configure(yscrollcommand=vscrollbar.set)
	canvas.pack()

	vscrollbar.config(command=canvas.yview)

	interieur=Frame(canvas,bg='white')
	canvas.create_window((0,0), window=interieur,anchor=NW,width=w-150,height=h-100)
	interieur.bind("<Configure>",myfunction)


	titre=Label(interieur, text="CARACTERISTIQUES DES VARIABLES",bg='white', fg="blue",font="times 15 bold")
	titre.pack(pady=15, side=TOP)

	panel_features1=PanedWindow(interieur,bg='white', orient=HORIZONTAL, width=w-160,height=50)
	panel_features1.pack(side=TOP)

	panel_features2=PanedWindow(interieur,bg='white', orient=HORIZONTAL, width=w-160,height=800) #, bg="#aaf033"
	panel_features2.pack(side=TOP, padx=40, pady=20)


	Label(panel_features1,bg='white',font="times 12 bold",text="Valeur manquante : ").pack(side=LEFT, padx=20)

	val_manq=StringVar()
	val_manq.set('0.00 %')
	Label(panel_features1, textvariable=val_manq, width=20,bg='white', fg="blue",font="times 12 bold").pack(side=LEFT)

	Label(panel_features1,bg='white',font="times 12 bold",text="Variable cible (à prédire) : ").pack(side=LEFT, padx=20)

	nb_lig=StringVar()
	nb_lig.set(type_target)
	Label(panel_features1, textvariable=nb_lig, bg='white',width=20, fg="blue",font="times 12 bold").pack(side=LEFT)
	#Label(lab_titre2, text=" lignes",width=5).pack(side=LEFT)

	frame_corr=Frame(panel_features2)
	frame_corr.pack(side=LEFT, padx=10)

	frame_hist=Frame(panel_features2)
	frame_hist.pack(side=LEFT, padx=10)
	
	fig1=Figure(figsize=(5,5),dpi=100)
	im1 = fig1.add_subplot(111)
	
	corr = data.corr()
	ax=sns.heatmap(corr,annot=True,vmin=-1, vmax=1,
					 center=0,
					cmap=sns.diverging_palette(20, 220, n=200),
					square=True, ax=im1)
	im1.set_xticklabels(
		im1.get_xticklabels(),
		rotation=45,
		horizontalalignment='right'
	);
	im1.set_yticklabels(
		im1.get_xticklabels(),
		rotation=0,
		horizontalalignment='right'
	);

	im1.set_title("Corrélation des variables")

	graph1 = FigureCanvasTkAgg(fig1, master=frame_corr)
	canvas1 = graph1.get_tk_widget()
	canvas1.pack(side=LEFT, padx=0)

	# HISTOGRAMME DE VARIABLE TARGET
	f = Figure(figsize=(5,5), dpi=100)

	p = f.add_subplot(111)
	p.hist(data[type_target], bins=20)
	#p.sns.distplot(data[type_target])
	p.set_title('Histogramme de variable cible')
	p.set_ylabel("Effectifs")
	p.set_xlabel("Valeur réelle")

	gqq = FigureCanvasTkAgg(f, master=frame_hist)
	canv2=gqq.get_tk_widget().grid()
	#canv2.pack(side=LEFT,padx=10)


	############################################################################
	###################   FIN INTERFACE  FEATURES ##############################
	############################################################################

def score():
	global score_moyenne

	dict_score_train={nom:valeur for nom,valeur in zip(nom_model,list_score_train)}
	dict_score_val={nom:valeur for nom,valeur in zip(nom_model,list_score_val)}

	score_moyenne=stq.mean(list_score_train).round(4)

	for i in range(len(list_score_train)):
		if list_score_train[i]<0:
			list_score_train[i]=0.5
			dict_score_val={nom:valeur for nom,valeur in zip(nom_model,list_score_train)}     

	for i in range(len(list_score_val)):
		if list_score_val[i]<0:
			list_score_val[i]=0.5
			dict_score_val={nom:valeur for nom,valeur in zip(nom_model,list_score_val)}

	#print("Validation finale : ",dict_score_val)

	index_plot=list(dict_score_train.keys())
	bar1=list(dict_score_train.values())
	bar2=list(dict_score_val.values())

	def affichage_score():
		fig_arbre=Figure(figsize=(12,5),dpi=100)
		width = 0.3
		x = np.arange(len(index_plot))

		im_arbre = fig_arbre.add_subplot(111)
		im_arbre.bar(x - width/2, bar1,color='b',width = 0.3, label="Entrainement")
		im_arbre.bar(x + width/2, bar2,color='r',width = 0.3, label="Validation") # marron, color ='b', width = 0.5
		im_arbre.set_xticks(x)
		im_arbre.set_xticklabels(index_plot)
		im_arbre.set_title("Comparaison de score")
		im_arbre.set_xlabel("Algorithmes")
		im_arbre.set_ylabel("Performance")
		im_arbre.legend()
		graph_arbre = FigureCanvasTkAgg(fig_arbre, master=panel_score1)
		canvas_arbre = graph_arbre.get_tk_widget()
		canvas_arbre.pack(padx=0)

	#dic = OrderedDict(sorted(dict_model.items(), key=lambda x: x[1],reverse=True))

	#cm=metrics.confusion_matrix(y_train,prediction)

	############################################################################
	###################   DEBUT INTERFACE  MENU SCORE  #######################
	############################################################################


	def selection_model(event=None):
		global date_prev
		if event:
			date_prev=event.widget.get()


	def myfunction(event):
		canvas.configure(scrollregion=canvas.bbox("all"),width=w-150,height=h-100)

	frame_score.pack(side=TOP, padx=0)

	mb.configure(state='normal')
	frame_features.pack_forget()
	frame_data.pack_forget()
	frame_general.pack_forget()
	frame_model.pack_forget()
	frame_prevision.pack_forget()
	frame_accueil.pack_forget()
	frame_historique.pack_forget()


	# create a canvas object and a vertical scrollbar for scrolling it
	vscrollbar = Scrollbar(frame_score, orient=VERTICAL)
	vscrollbar.pack(fill=Y, side=RIGHT, expand=FALSE)

	canvas = Canvas(frame_score, bd=0, highlightthickness=0)
	canvas.configure(yscrollcommand=vscrollbar.set)
	canvas.pack()

	vscrollbar.config(command=canvas.yview)

	interieur=Frame(canvas,bg='white')
	canvas.create_window((0,0), window=interieur,anchor=NW,width=w-150,height=h+300)
	interieur.bind("<Configure>",myfunction)


	titre=Label(interieur, bg='white',text="COMPARAISON DES SCORES D'ENTRAINEMENT ET DE VALIDATION", fg="blue",font="times 15 bold")
	titre.pack(pady=20, side=TOP)

	panel_score1=PanedWindow(interieur,bg='white', orient=HORIZONTAL, width=w-160,height=600)
	panel_score1.pack(side=TOP)

	panel_score2=PanedWindow(interieur,bg='white', orient=HORIZONTAL, width=w-160,height=800) #, bg="#aaf033"
	panel_score2.pack(side=TOP, padx=40, pady=20)

	frm_score=Frame(panel_score2)
	frm_score.pack(side=LEFT, padx=10)
	frm_confusion=Frame(panel_score2)
	frm_confusion.pack(side=LEFT, padx=10)


	affichage_score()

	def matrice_confusion():

		fig1=Figure(figsize=(5,5),dpi=100)
		im1 = fig1.add_subplot(111)

		ax=sns.heatmap(mat_conf,annot=True, ax=im1)

		im1.set_title("Matrice de confusion")
		im1.set_xlabel("Prédiction")
		im1.set_ylabel("Vraie valeur")

		graph1 = FigureCanvasTkAgg(fig1, master=frm_confusion)
		canvas1 = graph1.get_tk_widget()
		canvas1.pack(side=LEFT, padx=0)

	if type_dataset==liste_dataset[0]:
		if type_entrainement==liste_entrainement[1]:
			matrice_confusion()

	############################################################################
	###################   DEBUT INTERFACE  MENU SCORE ########################
	############################################################################
def historique():
	############################################################################
	###################   DEBUT INTERFACE  MENU HISTORIQUE  ####################
	############################################################################

	if type_dataset==liste_dataset[0]:
		if type_entrainement==liste_entrainement[0]:
			type_train=liste_entrainement[0]
		else:
			type_train=liste_entrainement[1]
	else:
		type_train=liste_entrainement[0]

	fic=str(name_fic)
	actuel=str(datetime.datetime.today().strftime('%d-%m-%Y  %H:%M:%S'))
	data=str(type_dataset)
	train=str(type_train)
	scr=str(score_moyenne)
	nb_mod=str(nb_model)
	targ=str(type_target)
	val_pred=str(pred)
	#x = str(x).encode('utf-8','ignore')

	liste_historique=[fic,actuel,data,train,scr,nb_mod,targ,val_pred]  

	requete='''INSERT into historique(Nom,Date_Creation , Type_data , Type_traitement , Score , Modele , Target ,Prediction ) VALUES(?,?,?,?,?,?,?,?)'''
	c.execute(requete, (fic,actuel,data,train,scr,nb_mod,targ,val_pred))
	conn.commit()

	def View():
		c.execute("SELECT * FROM historique")
		rows = c.fetchall()
		for row in rows:
			tableau.insert("", END, values=row)

	def exporter():
		import sqlite3
		from xlsxwriter.workbook import Workbook
		from xlrd import open_workbook
		from openpyxl import load_workbook
		workbook = Workbook('output.xlsx')
		worksheet = workbook.add_worksheet()

		conn=sqlite3.connect('prevision.db')
		c=conn.cursor()
		#c.execute("select * from abc")
		mysel=c.execute("select * from historique ")
		for i, row in enumerate(mysel):
			for j, value in enumerate(row):
				worksheet.write(i, j, value)
		#load_workbook(workbook)
		workbook.close()

	def supprimer_selected():
		selected_item = tableau.selection()[0]
		tableau.delete(selected_item)

	def selection_model(event=None):
		global date_prev
		if event:
			date_prev=event.widget.get()


	def myfunction(event):
		canvas.configure(scrollregion=canvas.bbox("all"),width=w-150,height=h-100)

	frame_score.pack_forget()
	mb.configure(state='normal')
	frame_features.pack_forget()
	frame_data.pack_forget()
	frame_general.pack_forget()
	frame_model.pack_forget()
	frame_prevision.pack_forget()
	frame_accueil.pack_forget()
	frame_historique.pack(side=LEFT, padx=0)


	# create a canvas object and a vertical scrollbar for scrolling it
	vscrollbar = Scrollbar(frame_historique, orient=VERTICAL)
	vscrollbar.pack(fill=Y, side=RIGHT, expand=FALSE)

	canvas = Canvas(frame_historique, bd=0, highlightthickness=0)
	canvas.configure(yscrollcommand=vscrollbar.set)
	canvas.pack()

	vscrollbar.config(command=canvas.yview)

	interieur=Frame(canvas,bg='white')
	canvas.create_window((0,0), window=interieur,anchor=NW,width=w-150,height=h-100)
	interieur.bind("<Configure>",myfunction)


	titre=Label(interieur,bg='white', text="TABLEAU RECAPUTILATIF DES HISTORIQUES", fg="blue",font="times 15 bold")
	titre.pack(pady=20, side=TOP)

	panel_score1=PanedWindow(interieur,bg='white', orient=HORIZONTAL, width=w-160,height=50)
	panel_score1.pack(side=TOP)

	panel_score2=PanedWindow(interieur,bg='white', orient=HORIZONTAL, width=w-160,height=900) #, bg="#aaf033"
	panel_score2.pack(side=TOP, padx=40, pady=20)

	lab_titre1 = LabelFrame(panel_score1,bg='white', text="Actualisation",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	lab_titre1.pack(expand="yes",padx=40, side=LEFT)

	Button(lab_titre1, text='Actualiser', width=20, bg='green', fg='white', command=View).pack()

	lab_titre = LabelFrame(panel_score1,bg='white', text="Suppression",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	lab_titre.pack(expand="yes",padx=40, side=LEFT)
	Button(lab_titre,text='Supprimer', width=20, bg='green', fg='white', command=supprimer_selected).pack()

	lab_titre2 = LabelFrame(panel_score1,bg='white', text="Triage",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	lab_titre2.pack(expand="yes",padx=40, side=LEFT)
	Button(lab_titre2,text='Trier', width=20, bg='green', fg='white').pack()


	lab_titre3 = LabelFrame(panel_score1,bg='white', text="Export vers Excel",width=(w-160)/4, padx=20, pady=20, relief=RIDGE)
	lab_titre3.pack(expand="yes",padx=40, side=LEFT)
	Button(lab_titre3,text='Exporter', width=20, bg='green', fg='white', command=exporter).pack()

	#Tableau affichage d'historique

	scrollbary = Scrollbar(panel_score2, orient=VERTICAL)

	tableau = ttk.Treeview(panel_score2,selectmode="browse",height=400,yscrollcommand=scrollbary.set)
	scrollbary.config(command=tableau.yview)
	scrollbary.pack(side=RIGHT, fill=Y)

	tableau["columns"]=("nom", "date", "type_data","type_traitement",
				"score", "modele","target", "prediction")
	tableau['show'] = 'headings'

	tableau.column('nom', width=130)
	tableau.heading('nom', text='Nom du fichier')

	tableau.column('date', width=130,anchor="center")
	tableau.heading('date', text='Date de prévision')

	tableau.column('type_data', width=130 )
	tableau.heading('type_data', text='Type de data')

	tableau.column('type_traitement',width=130)
	tableau.heading('type_traitement', text="Type d'entrainement")

	tableau.column('score', width=130, anchor="center")
	tableau.heading('score', text='Score')

	tableau.column('modele', width=130, anchor="center")
	tableau.heading('modele', text='Nombre de modele')

	tableau.column('target', width=130)
	tableau.heading('target', text='Variable cible')

	tableau.column('prediction', width=130, anchor="center")
	tableau.heading('prediction', text='Prédiction')

	tableau.pack(padx = 10, pady = (10, 10))

	#Affichage dans le tableau

	c.execute("SELECT * FROM historique")
	rows = c.fetchall()
	for row in rows:
		tableau.insert("", END, values=row)
	#conn.close()


	############################################################################
	###################     FIN INTERFACE  MENU HISTORIQUE #####################
	############################################################################
def creer_compte():
	def register():

		user=str(nom_societe.get())
		pwd1=str(pwd_societe1.get())
		pwd2=str(pwd_societe2.get())
		if pwd1==pwd2:
			requete='''INSERT into login(username,password) VALUES(?,?)'''
			c.execute(requete, (user,pwd1))
			conn.commit()
			showinfo('Information','Utilisateur enregistré!')
		else:
			showerror('Incorrect','Mot de passe incorrect!')

		profile.destroy()

	profile=Toplevel(fenetre, width=200, height=300)

	profile.title("Création de compte")
	
	nom_societe=StringVar()
	pwd_societe1=StringVar()
	pwd_societe2=StringVar()

	lab_user=Label(profile, text="Nom d'utilisateur :")
	lab_user.grid(row=0, column=0)
	user=Entry(profile, width=20, textvariable=nom_societe, font='times 14')
	user.grid(row=0, column=1)
	lab_mdp=Label(profile, text="Votre mot de passe :")
	lab_mdp.grid(row=1, column=0)
	pwd=Entry(profile, width=20, textvariable=pwd_societe1 , font='times 14')
	pwd.grid(row=1, column=1)

	lab_mdp=Label(profile, text="Confirmer le mot de passe ") 
	lab_mdp.grid(row=2, column=0)
	pwd=Entry(profile, width=20, textvariable=pwd_societe2 , font='times 14')
	pwd.grid(row=2, column=1)
	
	quit=Button(profile, width=20,text="Annuler",bg="green",fg="white", command=profile.destroy)
	quit.grid(row=4, column=0)
	btn_modif=Button(profile, width=25,text="Enregistrer",bg="green",fg="white", command=register)
	btn_modif.grid(row=4, column=1)
#menu
menubar = Menu(fenetre)

menu1 = Menu(menubar, tearoff=0)
menu1.add_command(label="Nouveau", command=importer)
menu1.add_command(label="Enregistrer")
menu1.add_command(label="Enregistrer sous...")
menu1.add_separator()
menu1.add_command(label="Quitter", command=fenetre.quit)
menubar.add_cascade(label="Fichier", menu=menu1)

menu2 = Menu(menubar, tearoff=0)
menu2.add_command(label="Exécuter", command=executer)
menu2.add_command(label="Pauser")
menu2.add_command(label="Arreter")
menubar.add_cascade(label="Editer", menu=menu2)

menu3 = Menu(menubar, tearoff=0)
menu3.add_command(label="A propos")
menu3.add_command(label="Créer un compte", command=creer_compte)
menubar.add_cascade(label="Aides", menu=menu3)


#panel general
p = PanedWindow(fenetre, orient=VERTICAL,relief=RIDGE)
p.pack(side=TOP, expand=Y, fill=BOTH)

#panel en tete
pane_tete=PanedWindow(p, orient=VERTICAL, width=w, height=40,relief=RIDGE)
pane_tete.pack(side=TOP, expand=Y,fill=BOTH)

#panel sidebar+contenu
pane_sidebar_contenu=PanedWindow(p, orient=HORIZONTAL, width=w, height=h,relief=RIDGE)
pane_sidebar_contenu.pack(side=RIGHT, expand=Y,fill=BOTH)

#panel en sidebar
pane_sidebar=PanedWindow(pane_sidebar_contenu, orient=VERTICAL,bg='#4e73df', width=160, height=h,relief=RIDGE)
pane_sidebar.pack(side=LEFT)


#panel en contenu
pane_contenu=PanedWindow(pane_sidebar_contenu, orient=VERTICAL, bg='white',width=w-160, height=h,relief=FLAT)
pane_contenu.pack(side=TOP, padx=0)


####################################En tete ###############################

# frame 1=entete
en_tete = Frame(pane_tete, borderwidth=3, relief=FLAT, bg='#4e73df')
en_tete.grid(row=0, column=0, sticky='N')
en_tete.place(width=w,height=40 )

#panel d'en tete

panel_tete = PanedWindow(en_tete, orient=HORIZONTAL, bg='#4e73df')
panel_tete.pack(side=LEFT, expand=Y, fill=BOTH)

import time

###### FONCTION D'AFFICHAGE DE L'HEURE ######
def Heure():
	Label_Heure.config(text=time.strftime('%a %d %b %Y   %H:%M:%S')) #%a
	Label_Heure.after(200, Heure) 

panel_tete.add(Label(panel_tete,text="PREVISION DE VENTE",bg='#4e73df',fg="white",font="times 15 bold "))
Label_Heure=Label(width=22,bg='#4e73df', fg='white',font='times 13 bold ')
Label_Heure.pack()
panel_tete.add(Label_Heure)

Heure()
btn_predire=Button(text="/>  Prédire ", bg='green', fg='orange',font="times 13 bold ",width=10,relief=GROOVE, command=executer)
btn_predire.pack()
btn_predire.configure(state='disable')
panel_tete.add(btn_predire)

#panel_tete.add(Button(panel_tete,text="I>  Prédire", bg='green', fg='orange',font="times 13 bold ",width=10,relief=GROOVE, command=executer))

def btn_recherche(event):
	selected=recherche.get()
	
	chearch="SELECT * from historique WHERE selected=?"
	select=c.execute(chearch, selected)

def btn_reset(event):
	value.set("") 

def parametre():
	param=Toplevel(fenetre, width=200, height=200)

	param.title("Parametres")
	
	lab_frame_bg=LabelFrame(param,text='Parametre de background', width=200, height=100,fg="blue" ,font="sans 11 bold italic")
	lab_frame_bg.pack(pady=8)
	lab_frame_font=LabelFrame(param,text='Parametre de font', width=200, height=100,fg="blue", font="sans 11 bold italic")
	lab_frame_font.pack(pady=8)
	lab_frame_police=LabelFrame(param,text='Taille de police', width=200,fg="blue", height=100, font="sans 11 bold italic")
	lab_frame_police.pack(pady=8)

	frame_btn=Frame(param)
	frame_btn.pack(side=TOP)

	bg_default = Checkbutton(lab_frame_bg, text="Par défaut")
	bg_default.pack(side=LEFT, pady=20)
	bg_gris = Checkbutton(lab_frame_bg, text="Gris")
	bg_gris.pack(side=LEFT, padx=35)

	font_default = Checkbutton(lab_frame_font, text="Par défaut")
	font_default.pack(side=LEFT, pady=20)
	font_times = Checkbutton(lab_frame_font, text="Times")
	font_times.pack(side=LEFT, padx=30)

	spin=Spinbox(lab_frame_police,from_=8, to=15)
	spin.pack(pady=20, padx=35)

	btn_validation=Button(frame_btn,text='Valider', width=15, bg='green', fg='white', command=param.destroy)
	btn_validation.pack(side=LEFT)
	btn_annulation=Button(frame_btn,text='Annuler', width=15,bg='green', fg='white', command=param.destroy)
	btn_annulation.pack(side=LEFT)

def apropos():
	propos=Toplevel(fenetre, width=200, height=200)
	propos.title("A propos")

	Label(propos,fg="blue",text="PDV Standard Version 1.0.0 est une platforme de prévision de vente basée sur Machine Learning\n C'est un memoire de fin d'etude d'obtention de diplome de Master titre Ingénieur").pack()
	Label(propos,text="Domaine : Science d'Ingénieur\nMention : ELECTRONIQUE\nParcours : INFORMATIQUE APPLIQUEE").pack()
	Label(propos,fg='orange',font="times 12 bold",text="Titre :\n Modélisation et Programmation d'un outil de prévision de vente").pack()
	Label(propos,text="Dévéloppé et présenté par :\n Monsieur ANDRIATSITOHAINA Elie Fenohasina\nMonsieur RAKOTOMALALA Maheriniaina Jacklin").pack()
	
	fr=Frame(propos)
	fr.pack()

	photo_accueil = PhotoImage(file="accueil5.png")
	canvas1 = Canvas(fr,width=200, height=200)
	canvas1.pack(side=TOP)
	canvas1.create_image(100, 100, image=photo_accueil)
	
def accueil():
	frame_data.pack_forget()
	frame_general.pack_forget()
	frame_model.pack_forget()
	frame_prevision.pack_forget()
	frame_features.pack_forget()
	frame_score.pack_forget()
	frame_historique.pack_forget()

	frame_accueil.pack(padx=20, side=LEFT)
	frame_accueil.place(x=0,y=0)
	#showerror('Information',"Page d'accueil non disponible!")
	 
value = StringVar() 
value.set("Rechercher...")
recherche=Entry(textvariable=value,width=25,font='sans 10 ')
recherche.pack()

recherche.bind("<Button-1>", btn_reset)
recherche.bind("<Return>", btn_recherche)

panel_tete.add(recherche)
panel_tete.add(Button(panel_tete,text="ACCUEIL",width=15,font='times 11 bold ',fg='white',bg='#4e73df',relief=FLAT, command=accueil))
panel_tete.add(Button(panel_tete,text="PARAMETRES",width=15,font='times 11 bold ',fg='white',bg='#4e73df',relief=FLAT, command=parametre))
panel_tete.add(Button(panel_tete,text="A PROPOS",width=15,font='times 11 bold ',fg='white',bg='#4e73df',relief=FLAT, command=apropos))
panel_tete.add(Label(panel_tete,text="",width=15,bg='#4e73df',relief=FLAT))

societe=StringVar()
societe.set("PROFIL")
mb=  Menubutton ( panel_tete, textvariable=societe,width=10,font='times 11 bold ',fg='white', relief=FLAT ,bg='#4e73df')
mb.menu =  Menu ( mb, tearoff = 0 )
mb["menu"] =  mb.menu

def profile():   

	def modifier_profile():

		new_pwd=str(pwd_societe.get())

		sql = "UPDATE login SET password = ? WHERE username = ?"
		c.execute(sql,(new_pwd,user_name))
		conn.commit()

		showinfo('Information','Mot de passe modifié!')
		profile.destroy()

	profile=Toplevel(fenetre, width=200, height=200)

	profile.title("Profil")

	nom_societe=StringVar()
	pwd_societe=StringVar()

	requete_profile="SELECT * from login WHERE username=?"
	query=c.execute(requete_profile,(user_name,))
	for res in query:
		nom_societe.set(res[0])
		pwd_societe.set(res[1])
	
	lab_user=Label(profile, text="Utilisateur :")
	lab_user.grid(row=0, column=0)
	user=Entry(profile, width=20, textvariable=nom_societe, font='times 14', state='disable')
	user.grid(row=0, column=1)
	lab_mdp=Label(profile, text="Mot de passe :")
	lab_mdp.grid(row=1, column=0)
	pwd=Entry(profile, width=20, textvariable=pwd_societe , font='times 14')
	pwd.grid(row=1, column=1)
	
	quit=Button(profile, width=20,text="Annuler",bg="green",fg="white", command=profile.destroy)
	quit.grid(row=4, column=0)
	btn_modif=Button(profile, width=25,text="Modifier",bg="green",fg="white", command=modifier_profile)
	btn_modif.grid(row=4, column=1)

def deconnecter():
	result = askquestion("Déconnexion", "Confirmer la déconnexion?", icon='warning')
	if result == 'yes':
		conn.close()
		fenetre.destroy()

def version():
	showinfo('Version','Prévision De Vente Standard \n Version 1.0.0')


mb.menu.add_command ( label="Mon profil", command=profile)
mb.menu.add_separator()
mb.menu.add_command ( label="PDV v1.0.0",command=version)
mb.menu.add_separator()
mb.menu.add_command ( label="Déconnexion",command=deconnecter)


mb.pack(side=RIGHT, padx=20)
mb.configure(state='disable')
############################## SIDEBAR #########################################

# frame 2=sidebar
sidebar = Frame(pane_sidebar, borderwidth=4, relief=FLAT, bg='#4e73df')
sidebar.grid(row=0, column=0, sticky='N')
sidebar.place(width=159,height=h )

#panel de sidebar
panel_sidebar = PanedWindow(sidebar, orient=VERTICAL,bg='#4e73df')
panel_sidebar.pack(side=LEFT, expand=Y, fill=BOTH)

panel_sidebar.add(Button(panel_sidebar,text="Se connecter",height=2,width=12,font='Copperplate 10 bold',fg='white',bg='#4e73df',relief=GROOVE, command=login))
panel_sidebar.add(Button(panel_sidebar,text="Données",height=2,width=15,font='Copperplate 10 bold',fg='white',bg='#4e73df',relief=GROOVE, command=dataset))
panel_sidebar.add(Button(panel_sidebar,text="Générale",height=2,width=15,font='Copperplate 10 bold',fg='white',bg='#4e73df',relief=GROOVE, command=general))
panel_sidebar.add(Button(panel_sidebar,text="Modèles",height=2,width=15,font='Copperplate 10 bold',fg='white',bg='#4e73df',relief=GROOVE, command=model))
panel_sidebar.add(Button(panel_sidebar,text="Caracteristiques",height=2,width=15,font='Copperplate 10 bold',fg='white',bg='#4e73df',relief=GROOVE, command=features))
panel_sidebar.add(Button(panel_sidebar,text="Score",height=2,width=15,font='Copperplate 10 bold',fg='white',bg='#4e73df',relief=GROOVE, command=score))   
panel_sidebar.add(Button(panel_sidebar,text="Prévision",height=2,width=15,font='Copperplate 10 bold',fg='white',bg='#4e73df',relief=GROOVE, command=prevision))
panel_sidebar.add(Button(panel_sidebar,text="Journal",height=2,width=15,font='Copperplate 10 bold',fg='white',bg='#4e73df',relief=GROOVE, command=historique))
panel_sidebar.add(Label(panel_sidebar,text="",pady=h,height=10, width=15,font='Copperplate 10 bold',fg='white',bg='#4e73df'))

################################### CONTENU   ###################################

frame_accueil=Frame(pane_contenu, width=w-160, height=h)
frame_accueil.pack(padx=20, side=LEFT)
frame_accueil.place(x=0,y=0)

photo_accueil = PhotoImage(file="accueil5.png")

canvas = Canvas(frame_accueil,width=w-160, height=h)
canvas.create_image(0, 0, anchor=NW, image=photo_accueil)
#canvas.pack()

def myfunction(event):
	canvas.configure(scrollregion=canvas.bbox("all"),width=w-200,height=h-100)

myscrollbar=Scrollbar(frame_accueil,orient="vertical",command=canvas.yview)
canvas.configure(yscrollcommand=myscrollbar.set)

myscrollbar.pack(side="right",fill="y")
canvas.pack(side="left")
frame_accueil.bind("<Configure>",myfunction)


# cacher les frame de data, parametre
frame_data = Frame(pane_contenu, borderwidth=1)
frame_data.pack_forget()

frame_general=Frame(pane_contenu)
frame_general.pack_forget()

frame_model=Frame(pane_contenu)
frame_model.pack_forget()

frame_prevision=Frame(pane_contenu)
frame_prevision.pack_forget()

frame_features=Frame(pane_contenu)
frame_features.pack_forget()

frame_score=Frame(pane_contenu)
frame_score.pack_forget()

frame_historique=Frame(pane_contenu)
frame_historique.pack_forget()


########################################################################################
######################  FIN INTERFACE GENERALE #########################################
########################################################################################

fenetre.config(menu=menubar)
fenetre.mainloop()
